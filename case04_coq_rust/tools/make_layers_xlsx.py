#!/usr/bin/env python3
"""Case 04 — Excel table of Arm A's layered structure, alongside the TikZ
diagram. Data from classify_extracted.analyze() (same source as the
diagram and the artifact-map report). Output: figures/armA_layers.xlsx
with sheets: Summary, Layers, Definitions, Theorems."""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from classify_extracted import analyze, LAYERS  # noqa: E402

WS = "/home/xcos/fcdd_c04_ds/armA"
WSB = "/home/xcos/fcdd_c04_ds/armB"
ORDER = [l for l, _ in LAYERS] + ["other"]

# Arm B: flat hand-written Rust — map its functions onto the same layers
ARMB_FNS = {
    "byte/bit layer": [],
    "squares & pieces": ["sq_of", "rank_of", "file_of", "in_bounds",
                         "Color", "Piece", "Cell", "Position::new",
                         "Position::piece_at", "Position::king_sq"],
    "move encoding & dirs": ["Move", "parse_square", "sq_str", "promo_char",
                             "move_str"],
    "attacks": ["attacked_by_pawn", "attacked", "in_check"],
    "move generation": ["gen_pawn_moves", "gen_knight_moves", "gen_sliding",
                        "gen_king_moves", "gen_castling", "pseudo_legal"],
    "make/legality": ["push_if_legal", "apply_move", "legal_moves"],
    "terminal & draw": ["Status", "insufficient_material", "status_of"],
    "evaluation": [],
    "search & perft": [],
    "FEN / CLI": ["parse_fen", "parse_args", "main"],
    "other": [],
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
STRIPE = PatternFill("solid", fgColor="EAF1F8")


def sheet_of(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(c)].width = w
    return ws


def main():
    data = analyze(WS)
    layers, gen, theorems, defs = (data["layers"], data["gen"],
                                   data["theorems"], data["defs"])
    tset = {n: refs for n, refs in theorems}

    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary ---
    ws = sheet_of(wb, "Summary",
                  ["Metric", "Value"], [42, 60])
    total = sum(len(v) for v in gen.values())
    rows = [
        ("Rocq specification definitions (ChessSpec.v)", len(defs)),
        ("Definitions with generated Rust code", len(gen)),
        ("Definitions without generated code (types/constants)",
         ", ".join(sorted(set(defs) - set(gen)))),
        ("Generated Rust functions (extracted.rs)", total),
        ("  of which __curried wrappers",
         sum(1 for v in gen.values() if "curried" in v)),
        ("Kernel-checked theorems (Theorems.v)", len(theorems)),
        ("Layers in the model", sum(1 for l in ORDER if layers.get(l))),
        ("Binary size (release)", "547,616 bytes"),
    ]
    for r, (a, b) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)

    # --- Layers ---
    ws = sheet_of(wb, "Layers",
                  ["Layer", "Definitions", "Extracted", "Example definitions",
                   "Theorems constraining the layer"],
                  [26, 12, 10, 60, 60])
    r = 2
    for layer in ORDER:
        names = sorted(layers.get(layer, []))
        if not names:
            continue
        nd = len(names)
        ng = sum(1 for n in names if n in gen)
        ths = [t for t, refs in tset.items() if set(refs) & set(names)]
        ws.cell(row=r, column=1, value=layer).font = Font(bold=True)
        ws.cell(row=r, column=2, value=nd)
        ws.cell(row=r, column=3, value=ng)
        ws.cell(row=r, column=4,
                value=", ".join(names[:6]) + ("  (+%d more)" % (nd - 6) if nd > 6 else ""))
        ws.cell(row=r, column=5, value=", ".join(ths))
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = STRIPE
        r += 1

    # --- Definitions ---
    ws = sheet_of(wb, "Definitions",
                  ["Definition", "Kind", "Layer", "Generated",
                   "Theorems referencing it"],
                  [30, 12, 22, 24, 60])
    r = 2
    for layer in ORDER:
        for name in sorted(layers.get(layer, [])):
            kind = defs[name][0]
            kinds = gen.get(name, set())
            gtxt = ("plain + curried" if kinds == {"plain", "curried"}
                    else ("plain" if kinds == {"plain"} else
                          ("curried" if kinds == {"curried"} else
                           "NOT EXTRACTED")))
            ths = [t for t, refs in tset.items() if name in refs]
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=kind)
            ws.cell(row=r, column=3, value=layer)
            ws.cell(row=r, column=4, value=gtxt)
            ws.cell(row=r, column=5, value=", ".join(ths))
            if r % 2 == 0:
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = STRIPE
            r += 1

    # --- Theorems ---
    ws = sheet_of(wb, "Theorems",
                  ["Theorem", "Statement (first line)", "Definitions it constrains"],
                  [30, 60, 70])
    r = 2
    thm_bodies = {}
    # re-parse Theorems.v for statements
    from classify_extracted import def_blocks
    for kind, name, body in def_blocks(os.path.join(WS, "rocq", "Theorems.v")):
        if kind in ("Theorem", "Lemma"):
            stmt = " ".join(body.splitlines()[1:4]).strip()
            thm_bodies[name] = stmt[:110]
    for name, refs in theorems:
        layer_names = []
        for r_ in refs:
            for l in ORDER:
                if r_ in layers.get(l, []):
                    layer_names.append("%s (%s)" % (r_, l))
                    break
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=thm_bodies.get(name, ""))
        ws.cell(row=r, column=3, value="; ".join(layer_names))
        if r % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = STRIPE
        r += 1

    # --- Arm B ---
    ws = sheet_of(wb, "ArmB", ["Layer", "Functions", "Notes"], [26, 60, 60])
    r = 2
    for layer in ORDER:
        fns = ARMB_FNS.get(layer, [])
        if not fns:
            continue
        ws.cell(row=r, column=1, value=layer).font = Font(bold=True)
        ws.cell(row=r, column=2, value=", ".join(fns))
        if layer == "evaluation":
            ws.cell(row=r, column=3,
                    value="NOT IMPLEMENTED — choose was not scored in the "
                          "D11 calibration; this is the gap D12 addresses")
        elif layer == "search & perft":
            ws.cell(row=r, column=3,
                    value="NOT IMPLEMENTED — same reason (D12 fills it)")
        if r % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = STRIPE
        r += 1

    # --- Comparison ---
    ws = sheet_of(wb, "Comparison",
                  ["Layer", "Arm A defs (Rocq)", "Arm A generated fns",
                   "Arm B functions", "Arm A theorems on layer"],
                  [26, 18, 20, 20, 60])
    r = 2
    for layer in ORDER:
        names = sorted(layers.get(layer, []))
        bfns = ARMB_FNS.get(layer, [])
        if not names and not bfns:
            continue
        nd = len(names)
        ng = sum(1 for n in names if n in gen)
        ths = [t for t, refs in tset.items() if set(refs) & set(names)]
        ws.cell(row=r, column=1, value=layer).font = Font(bold=True)
        ws.cell(row=r, column=2, value=nd if nd else "-")
        ws.cell(row=r, column=3, value=ng if ng else "-")
        ws.cell(row=r, column=4, value=len(bfns) if bfns else "-")
        ws.cell(row=r, column=5, value=", ".join(ths))
        if r % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = STRIPE
        r += 1

    for ws_ in wb.worksheets:
        ws_.freeze_panes = "A2"
        for row in ws_.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    out = os.path.join(os.path.dirname(HERE), "figures", "armA_layers.xlsx")
    wb.save(out)
    print("written %s" % out)


if __name__ == "__main__":
    main()
