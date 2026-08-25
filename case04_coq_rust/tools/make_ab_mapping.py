#!/usr/bin/env python3
"""Case 04 — Excel mapping between Arm A (Rocq, extracted) and Arm B
(direct Rust), function-level, with the gaps called out. The mapping is
semantic (Arm B's 30 coarser functions vs Arm A's 161 definitions); each
Arm B function lists its Arm A semantic equivalents, and the Gaps sheet
lists what each arm has that the other lacks.

Output: figures/armA_armB_mapping.xlsx
Sheets: Mapping (B-fn -> A defs -> layer -> gap note), Gaps, Summary."""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
STRIPE = PatternFill("solid", fgColor="EAF1F8")
GAP_FILL = PatternFill("solid", fgColor="FFF2CC")

# Arm B function -> (layer, [Arm A semantic equivalents], gap note)
MAPPING = [
    # squares & pieces
    ("sq_of", "squares & pieces", ["sqFile", "sqRank", "sq64"], ""),
    ("rank_of", "squares & pieces", ["sqRank"], ""),
    ("file_of", "squares & pieces", ["sqFile"], ""),
    ("in_bounds", "squares & pieces", ["onBoard"], ""),
    ("Color", "squares & pieces", ["WHITE", "BLACK", "COLBIT"], ""),
    ("Piece", "squares & pieces", ["WP..BK", "TYPEMASK", "pcType", "pcCol"], ""),
    ("Cell", "squares & pieces", ["(board cells)", "bGet", "bSet"], ""),
    ("Position::new", "squares & pieces", ["emptyBoard", "startPos"], ""),
    ("Position::piece_at", "squares & pieces", ["bGet"], ""),
    ("Position::king_sq", "squares & pieces", ["(king squares in Position)"], ""),
    # move encoding
    ("Move", "move encoding & dirs", ["Move", "mvFrm", "mvDst", "mvFlag"], ""),
    ("parse_square", "move encoding & dirs", ["(FEN sq decode)", "sqFile", "sqRank"], ""),
    ("sq_str", "move encoding & dirs", ["(render)", "sqName"], ""),
    ("promo_char", "move encoding & dirs", ["mvPromo", "promoFlags"], ""),
    ("move_str", "move encoding & dirs", ["mvSpecial", "mvPromo", "(render)"], ""),
    # attacks
    ("attacked_by_pawn", "attacks", ["pawnAt", "isAttacked"], ""),
    ("attacked", "attacks", ["isAttacked", "scanHop", "scanSlide", "slideHit"], ""),
    ("in_check", "attacks", ["inCheckSide"], ""),
    # move generation
    ("gen_pawn_moves", "move generation", ["genPawnWhite", "genPawnBlack", "addPromos"], ""),
    ("gen_knight_moves", "move generation", ["genHops", "knightDirs"], ""),
    ("gen_sliding", "move generation", ["genRay", "genSlides", "bishopDirs", "rookDirs"], ""),
    ("gen_king_moves", "move generation", ["genHops", "kingDirs"], ""),
    ("gen_castling", "move generation", ["genCastling", "SP_OO", "SP_OOO"], ""),
    ("pseudo_legal", "move generation", ["genMoves"], ""),
    # make / legality
    ("push_if_legal", "make/legality", ["genLegal", "moverInCheck"], ""),
    ("apply_move", "make/legality", ["makeMove", "capSquare", "clrCastleSq", "SP_EP", "SP_DPUSH"], ""),
    ("legal_moves", "make/legality", ["genLegal"], ""),
    # terminal & draw
    ("Status", "terminal & draw", ["GSplay..GSdraw", "statusOf"], ""),
    ("insufficient_material", "terminal & draw", ["isInsufficient"], ""),
    ("status_of", "terminal & draw", ["statusOf", "updateTerminal", "countReps", "History"], ""),
    # FEN / CLI
    ("parse_fen", "FEN / CLI", ["parseBoard", "parseBoardAux", "takeField", "dropField", "nzeros", "pieceOfChar"], ""),
    ("parse_args", "FEN / CLI", ["(argv parse in adapter)"], ""),
    ("main", "FEN / CLI", ["entry", "render"], ""),
]

# gaps: what each arm has that the other lacks
GAPS = [
    ("Arm A only — byte/bit emulation",
     "w8, addB, negB, andB, orB, xorB, bitn, pow2 (17 defs)",
     "The formal model expresses the Z80's 8-bit wrapping arithmetic. "
     "Arm B uses native Rust integers and never needs this layer."),
    ("Arm A only — evaluation",
     "pst, pieceVal, gamePhase, eval, matingEval (16 defs)",
     "Arm A implemented the engine's evaluation even though `choose` was "
     "not scored in the D11 calibration. Arm B has NO evaluation."),
    ("Arm A only — search",
     "chooseMove, perft (2 defs)",
     "Arm A implemented depth-1 negamax + perft. Arm B has NO search: its "
     "`choose` is a stub. This is the gap D12 turns into the primary outcome."),
    ("Arm A only — machine-checked proofs",
     "Theorems.v: 26 kernel-checked theorems, zero Admitted",
     "Arm B has no proofs of any kind. These constrain the definitions "
     "above and are erased at extraction."),
    ("Arm A only — extraction provenance",
     "artifacts/chess_extracted.rs.out + hash-locked extracted.rs",
     "The shipped Rust is provably the mechanical image of the spec. "
     "Arm B's code is the only artifact."),
    ("Arm B only — (none)",
     "",
     "Arm B is semantically a subset of Arm A: every B function maps onto "
     "A definitions. The reverse is not true."),
]


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # --- Mapping sheet ---
    ws = wb.create_sheet("Mapping")
    headers = ["Arm B function (direct Rust)", "Layer",
               "Arm A semantic equivalents (Rocq, extracted)",
               "Gap note"]
    widths = [30, 22, 60, 60]
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(c)].width = w
    r = 2
    for fn, layer, equivs, note in MAPPING:
        ws.cell(row=r, column=1, value=fn)
        ws.cell(row=r, column=2, value=layer)
        ws.cell(row=r, column=3, value=", ".join(equivs))
        ws.cell(row=r, column=4, value=note)
        if r % 2 == 0:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = STRIPE
        r += 1

    # --- Gaps sheet ---
    ws = wb.create_sheet("Gaps")
    headers = ["Gap", "Definitions / artifacts", "What it means"]
    for c, (h, w) in enumerate(zip(headers, (34, 52, 70)), 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(c)].width = w
    r = 2
    for g, defs, meaning in GAPS:
        ws.cell(row=r, column=1, value=g).font = Font(bold=True)
        ws.cell(row=r, column=2, value=defs)
        ws.cell(row=r, column=3, value=meaning)
        for c in range(1, 4):
            ws.cell(row=r, column=c).fill = GAP_FILL
        r += 1

    # --- Summary ---
    ws = wb.create_sheet("Summary")
    rows = [
        ("Metric", "Arm A (formal/extract)", "Arm B (direct Rust)"),
        ("Hand-written lines", "1,097 (Rocq) + 96 (adapter)", "991 (single main.rs)"),
        ("Generated lines", "11,064", "0"),
        ("Definitions / functions", "161 defs -> 150 extracted (257 fns)", "30 fns"),
        ("Theorems", "26 (zero Admitted)", "0"),
        ("Binary", "547,616 B", "498,424 B"),
        ("Layers", "10 semantic layers", "flat, maps onto 8 of the 10"),
        ("Missing layers", "—", "byte/bit emulation, evaluation, search"),
        ("choose implemented", "yes (depth-1 negamax + eval)", "no (stub)"),
        ("mu1 vs model referee (calibration)", "0.00000", "0.00000"),
    ]
    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if r == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            if c == 1:
                cell.font = Font(bold=True)
        ws.column_dimensions["A"].width = 34
        ws.column_dimensions["B"].width = 44
        ws.column_dimensions["C"].width = 34

    for ws_ in wb.worksheets:
        ws_.freeze_panes = "A2"
        for row in ws_.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    out = os.path.join(os.path.dirname(HERE), "figures",
                       "armA_armB_mapping.xlsx")
    wb.save(out)
    print("written %s" % out)


if __name__ == "__main__":
    main()
